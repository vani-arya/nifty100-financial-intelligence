import os
import pandas as pd

from src.screener.engine import run_screener
from src.screener.composite_score import (
    calculate_composite_score,
    calculate_sector_relative_score
)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_bluechip",
    "turnaround_watch"
]

EXPORT_COLUMNS = [
    "company_id",
    "broad_sector",
    "year",
    "market_cap_crore",

    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    "debt_to_equity",
    "interest_coverage",

    "free_cash_flow_cr",
    "cash_from_operations_cr",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",

    "composite_quality_score",
    "sector_relative_score"
]


def export_screeners():

    os.makedirs(
        "output",
        exist_ok=True
    )

    output_file = (
        "output/screener_output.xlsx"
    )

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for preset in PRESETS:

            print(
                f"Exporting {preset}"
            )

            df = run_screener(
                preset
            )

            if len(df) == 0:

                empty_df = pd.DataFrame(
                    {
                        "message": [
                            "No companies matched this screener."
                        ]
                    }
                )

                empty_df.to_excel(
                    writer,
                    sheet_name=preset[:31],
                    index=False
                )

                continue

            df = calculate_composite_score(
                df
            )

            df = calculate_sector_relative_score(
                df
            )

            df = df.sort_values(
                "composite_quality_score",
                ascending=False
            )

            export_df = df[
                EXPORT_COLUMNS
            ].copy()

            export_df.to_excel(
                writer,
                sheet_name=preset[:31],
                index=False
            )

    apply_colours(
        output_file
    )

    print(
        f"\nExport completed: {output_file}"
    )


def apply_colours(
    file_path
):

    workbook = load_workbook(
        file_path
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE"
    )

    for sheet in workbook.worksheets:

        headers = [
            cell.value
            for cell in sheet[1]
        ]

        for row in sheet.iter_rows(
            min_row=2
        ):

            for cell in row:

                column_name = headers[
                    cell.column - 1
                ]

                value = cell.value

                if value is None:
                    continue

                try:

                    if column_name == "return_on_equity_pct":
                        cell.fill = (
                            green_fill
                            if value > 15
                            else red_fill
                        )

                    elif column_name == "debt_to_equity":
                        cell.fill = (
                            green_fill
                            if value < 1
                            else red_fill
                        )

                    elif column_name == "free_cash_flow_cr":
                        cell.fill = (
                            green_fill
                            if value > 0
                            else red_fill
                        )

                    elif column_name == "revenue_cagr_5yr":
                        cell.fill = (
                            green_fill
                            if value > 10
                            else red_fill
                        )

                    elif column_name == "pat_cagr_5yr":
                        cell.fill = (
                            green_fill
                            if value > 15
                            else red_fill
                        )

                    elif column_name == "dividend_yield_pct":
                        cell.fill = (
                            green_fill
                            if value > 2
                            else red_fill
                        )

                except Exception:
                    pass

    workbook.save(
        file_path
    )


if __name__ == "__main__":

    export_screeners()