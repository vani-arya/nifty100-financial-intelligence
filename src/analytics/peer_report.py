import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill


GREEN_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

YELLOW_FILL = PatternFill(
    start_color="FFEB9C",
    end_color="FFEB9C",
    fill_type="solid"
)

RED_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

BENCHMARK_FILL = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid"
)


def load_peer_groups():

    return pd.read_excel(
        "data/raw/peer_groups.xlsx"
    )


def load_company_names():

    conn = sqlite3.connect(
        "data/nifty100.db"
    )

    df = pd.read_sql(
        """
        SELECT
            id,
            company_name
        FROM companies
        """,
        conn
    )

    conn.close()

    return df


def load_financial_data():

    conn = sqlite3.connect(
        "data/nifty100.db"
    )

    df = pd.read_sql(
        """
        SELECT *
        FROM financial_ratios
        """,
        conn
    )

    conn.close()

    return df


def get_latest_year_data(df):

    temp = df.copy()

    temp["numeric_year"] = (
        temp["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0]
        .astype(float)
    )

    latest = (
        temp.groupby("company_id")
        ["numeric_year"]
        .transform("max")
    )

    temp = temp[
        temp["numeric_year"] == latest
    ]

    temp = (
        temp.sort_values("id")
        .drop_duplicates(
            subset=["company_id"],
            keep="last"
        )
    )

    return temp


def build_dataset():

    peer_df = load_peer_groups()

    financial_df = load_financial_data()

    financial_df = get_latest_year_data(
        financial_df
    )

    company_df = load_company_names()

    merged = (
        peer_df.merge(
            company_df,
            left_on="company_id",
            right_on="id",
            how="left"
        )
        .merge(
            financial_df,
            on="company_id",
            how="left"
        )
    )

    return merged


def create_peer_report():

    df = build_dataset()

    metrics = [
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "return_on_equity_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "capex_cr",
        "earnings_per_share",
        "book_value_per_share",
        "dividend_payout_ratio_pct",
        "total_debt_cr",
        "cash_from_operations_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "composite_quality_score",
        "return_on_capital_employed_pct"
    ]

    wb = Workbook()

    wb.remove(
        wb.active
    )

    groups = sorted(
        df["peer_group_name"]
        .dropna()
        .unique()
    )

    for group in groups:

        ws = wb.create_sheet(
            title=group[:31]
        )

        group_df = df[
            df["peer_group_name"]
            == group
        ].copy()

        headers = [
            "company_id",
            "company_name"
        ]

        for metric in metrics:

            headers.append(metric)
            headers.append(
                f"{metric}_pct"
            )

        for col_num, header in enumerate(
            headers,
            start=1
        ):
            ws.cell(
                row=1,
                column=col_num
            ).value = header

        for metric in metrics:

            ascending = (
                metric
                != "debt_to_equity"
            )

            group_df[
                f"{metric}_pct"
            ] = (
                group_df[metric]
                .rank(
                    pct=True,
                    ascending=ascending
                )
            )

            if metric == "debt_to_equity":

                group_df[
                    f"{metric}_pct"
                ] = (
                    1
                    -
                    group_df[
                        f"{metric}_pct"
                    ]
                )

        row_num = 2

        for _, row in group_df.iterrows():

            values = [
                row["company_id"],
                row["company_name"]
            ]

            for metric in metrics:

                values.append(
                    row[metric]
                )

                values.append(
                    row[
                        f"{metric}_pct"
                    ]
                )

            for col_num, value in enumerate(
                values,
                start=1
            ):
                ws.cell(
                    row=row_num,
                    column=col_num
                ).value = value

            if row["is_benchmark"]:

                for cell in ws[row_num]:

                    cell.fill = (
                        BENCHMARK_FILL
                    )

            row_num += 1

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(
            2,
            max_row + 1
        ):

            for c in range(
                4,
                max_col + 1,
                2
            ):

                cell = ws.cell(
                    row=r,
                    column=c
                )

                value = cell.value

                if value is None:
                    continue

                if value >= 0.75:

                    cell.fill = (
                        GREEN_FILL
                    )

                elif value <= 0.25:

                    cell.fill = (
                        RED_FILL
                    )

                else:

                    cell.fill = (
                        YELLOW_FILL
                    )

        median_row = (
            ws.max_row + 2
        )

        ws.cell(
            row=median_row,
            column=1
        ).value = (
            "PEER GROUP MEDIAN"
        )

        col_index = 3

        for metric in metrics:

            ws.cell(
                row=median_row,
                column=col_index
            ).value = (
                group_df[
                    metric
                ]
                .median()
            )

            col_index += 2

    wb.save(
        "output/peer_comparison.xlsx"
    )

    print(
        "peer_comparison.xlsx created"
    )